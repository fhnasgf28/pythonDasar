from odoo import http
from odoo.http import request
from odoo.exceptions import ValidationError
import base64
import csv
import os
import re
import json
from io import BytesIO, TextIOWrapper
from openpyxl import load_workbook

class RFQImportPortal(http.Controller):
    @http.route(['/my/rfq/import_rfq_document'], type='json', auth="user", website=True)
    def portal_vendor_import_rfq_doc(self, file, file_name, name_sheet, **kw):
        fileformat = os.path.splitext(file_name)[1]
        data = []
        result = {
            'success': [],
            'failed': []
        }

        try:
            if fileformat == ".csv":
                decoded_datas = base64.b64decode(file)
                file = TextIOWrapper(BytesIO(decoded_datas))
                sniffer = csv.Sniffer()
                sniffer.preferred = [';', ',']
                dialect = sniffer.sniff(file.read())
                file.seek(0)
                csvreader = csv.reader(file)
                fields = next(csvreader)
                for row in csvreader:
                    data.append(row)
                result = self.rfq_update_import(data)
            elif fileformat == ".xls" or fileformat == ".xlsx":
                xlDecoded = base64.b64decode(file)
                workbook = load_workbook(filename=BytesIO(xlDecoded))
                data = []
                for sheet in workbook.worksheets:
                    if name_sheet and sheet.title != name_sheet:
                        continue
                    # Get header row to identify columns
                    headers = [cell.value for cell in sheet[1]]
                    
                    # Find the indices of required columns based on export format
                    rfq_name_idx = next((i for i, h in enumerate(headers) if h and h == 'Name'), None)
                    if rfq_name_idx is None:
                        rfq_name_idx = next((i for i, h in enumerate(headers) if h and ('rfq' in str(h).lower() or 'reference' in str(h).lower()) and ('name' in str(h).lower() or 'number' in str(h).lower())), None)
                    
                    product_idx = next((i for i, h in enumerate(headers) if h and h == 'Product'), None)
                    if product_idx is None:
                        product_idx = next((i for i, h in enumerate(headers) if h and 'product' in str(h).lower()), None)
                    qty_idx = next((i for i, h in enumerate(headers) if h and h == 'Quantity'), None)
                    if qty_idx is None:
                        qty_idx = next((i for i, h in enumerate(headers) if h and ('qty' in str(h).lower() or 'quantity' in str(h).lower())), None)

                    price_idx = next((i for i, h in enumerate(headers) if h and h == 'Price Unit'), None)
                    if price_idx is None:
                        price_idx = next((i for i, h in enumerate(headers) if h and ('price' in str(h).lower() or 'unit price' in str(h).lower())), None)
                    
                    if not all([rfq_name_idx is not None, qty_idx is not None, price_idx is not None, product_idx is not None]):
                        return {
                            'message': "Import Failed! Required columns missing. Please ensure your file has columns for RFQ Number, Product, Quantity, and Unit Price."
                        }
                    
                    # Process data rows
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                        if not row[rfq_name_idx]:  # Skip rows without RFQ name
                            continue
                        data.append({
                            'rfq_name': row[rfq_name_idx],
                            'product': row[product_idx],
                            'quantity': row[qty_idx],
                            'price_unit': row[price_idx],
                            'row': row_idx
                        })
                result = self.rfq_update_import(data)
            else:
                return {'message': "Please select xls/xlsx or csv file and try again!"}
                
            # Format the result message
            success_count = len(result['success'])
            failed_count = len(result['failed'])
            
            message = f"Import completed: {success_count} records updated successfully."
            if failed_count > 0:
                message += f"\n{failed_count} records failed to update."
                message += "\nFailed records:"
                for fail in result['failed']:
                    message += f"\n- Row {fail.get('row', 'N/A')}: {fail.get('message', 'Unknown error')}"
            
            return {'message': message}
        except Exception as e:
            return {'message': f"Import Failed!\nError: {str(e)}\nPlease check your file and try again"}

    def rfq_update_import(self, data):
        result = {
            'success': [],
            'failed': []
        }
        
        partner_id = request.env.user.partner_id.id
        
        for item in data:
            try:
                # For Excel import
                if isinstance(item, dict):
                    rfq_name = item['rfq_name']
                    product = item['product']
                    quantity = item['quantity']
                    price_unit = item['price_unit']
                    row = item['row']
                # For CSV import
                else:
                    rfq_name = item[0]
                    product = item[1]
                    quantity = item[2]
                    price_unit = item[3]
                    row = "N/A"
                
                # Validate data types
                try:
                    quantity = float(quantity)
                    price_unit = float(price_unit)
                except (ValueError, TypeError):
                    result['failed'].append({
                        'row': row,
                        'message': f"Invalid data format. Quantity and Unit Price must be numbers."
                    })
                    continue
                
                # Apply multi-company filtering if needed
                domain = [('name', '=', rfq_name), ('partner_id', '=', partner_id)]
                if request.env.context.get('allowed_company_ids'):
                    if len(request.env.companies) == 1:
                        domain.append(('company_id', '=', request.env.company.id))
                    else:
                        domain.append(('company_id', 'in', request.env.companies.ids))
                
                # Find the RFQ by name
                purchase_order = request.env['purchase.order'].sudo().search(domain, limit=1)
                
                if not purchase_order:
                    result['failed'].append({
                        'row': row,
                        'message': f"RFQ with reference '{rfq_name}' not found or you don't have access to it."
                    })
                    continue
                
                # Find the product
                product_domain = []
                # Check if product is a number (ID) or string (name/reference)
                try:
                    product_id = int(product)
                    product_domain = [('id', '=', product_id)]
                except (ValueError, TypeError):
                    # Try to find by name or default_code (reference)
                    product_domain = ['|', ('name', 'ilike', product), ('default_code', '=', product)]
                
                product_obj = request.env['product.product'].sudo().search(product_domain, limit=1)
                
                if not product_obj:
                    result['failed'].append({
                        'row': row,
                        'message': f"Product '{product}' not found."
                    })
                    continue
                
                # Find the order line with the specified product
                order_line = purchase_order.order_line.filtered(lambda l: l.product_id.id == product_obj.id)
                
                if not order_line:
                    result['failed'].append({
                        'row': row,
                        'message': f"Product '{product}' not found in RFQ '{rfq_name}'."
                    })
                    continue
                
                # Update quantity and unit price
                order_line.write({
                    'product_qty': quantity,
                    'price_unit': price_unit
                })
                
                result['success'].append({
                    'row': row,
                    'rfq_name': rfq_name,
                    'product': product
                })
                
            except Exception as e:
                result['failed'].append({
                    'row': row if 'row' in locals() else "N/A",
                    'message': str(e)
                })
        
        return result