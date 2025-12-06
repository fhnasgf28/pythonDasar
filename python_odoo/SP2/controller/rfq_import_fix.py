from odoo import http
from odoo.http import request
from odoo.exceptions import ValidationError
import base64
import csv
import os
import re
import json
from io import BytesIO, TextIOWrapper
from openpyxl import load_workbook


class RFQImportPortal(http.Controller):
    @http.route(['/my/rfq/import_rfq_document'], type='json', auth="user", website=True)
    def portal_vendor_import_rfq_doc(self, file, file_name, name_sheet, **kw):
        """Import RFQ data from Excel or CSV file"""
        fileformat = os.path.splitext(file_name)[1]
        data = []
        result = {
            'success': [],
            'failed': []
        }

        try:
            if fileformat == ".csv":
                decoded_datas = base64.b64decode(file)
                file = TextIOWrapper(BytesIO(decoded_datas))
                sniffer = csv.Sniffer()
                sniffer.preferred = [';', ',']
                dialect = sniffer.sniff(file.read())
                file.seek(0)
                csvreader = csv.reader(file)
                headers = next(csvreader)
                
                # Find the indices of required columns
                rfq_name_idx = self._find_column_index(headers, 'Name', ['rfq', 'reference', 'name', 'number'])
                product_idx = self._find_column_index(headers, 'Product', ['product'])
                qty_idx = self._find_column_index(headers, 'Quantity', ['qty', 'quantity'])
                price_idx = self._find_column_index(headers, 'Price Unit', ['price', 'unit price'])
                
                if not all([rfq_name_idx is not None, product_idx is not None, qty_idx is not None, price_idx is not None]):
                    return {
                        'message': "Import Failed! Required columns missing. Please ensure your file has columns for Name (RFQ Reference), Product, Quantity, and Price Unit."
                    }
                
                # Process data rows
                current_rfq = None
                for row_idx, row in enumerate(csvreader, 2):
                    if not row or len(row) <= max(rfq_name_idx, product_idx, qty_idx, price_idx):
                        continue
                    
                    # Handle empty RFQ ID (continuation of previous RFQ)
                    rfq_id = row[rfq_name_idx] if rfq_name_idx < len(row) and row[rfq_name_idx] else current_rfq
                    if not rfq_id:
                        continue
                    
                    current_rfq = rfq_id
                    
                    data.append({
                        'rfq_id': rfq_id,
                        'product_id': row[product_idx] if product_idx < len(row) and row[product_idx] else None,
                        'quantity': row[qty_idx] if qty_idx < len(row) and row[qty_idx] else None,
                        'price_unit': row[price_idx] if price_idx < len(row) and row[price_idx] else None,
                        'row': row_idx
                    })
                    
            elif fileformat in (".xls", ".xlsx"):
                xlDecoded = base64.b64decode(file)
                workbook = load_workbook(filename=BytesIO(xlDecoded))
                
                for sheet in workbook.worksheets:
                    if name_sheet and sheet.title != name_sheet:
                        continue
                        
                    # Get header row to identify columns
                    headers = [cell.value for cell in sheet[1]]
                    
                    # Find the indices of required columns
                    rfq_name_idx = self._find_column_index(headers, 'Name', ['rfq', 'reference', 'name', 'number'])
                    product_idx = self._find_column_index(headers, 'Product', ['product'])
                    qty_idx = self._find_column_index(headers, 'Quantity', ['qty', 'quantity'])
                    price_idx = self._find_column_index(headers, 'Price Unit', ['price', 'unit price'])
                    
                    if not all([rfq_name_idx is not None, product_idx is not None, qty_idx is not None, price_idx is not None]):
                        return {
                            'message': "Import Failed! Required columns missing. Please ensure your file has columns for Name (RFQ Reference), Product, Quantity, and Price Unit."
                        }
                    
                    # Process data rows
                    current_rfq = None
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                        if not row or len(row) <= max(rfq_name_idx, product_idx, qty_idx, price_idx):
                            continue
                        
                        # Handle empty RFQ ID (continuation of previous RFQ)
                        rfq_id = row[rfq_name_idx] if rfq_name_idx < len(row) and row[rfq_name_idx] else current_rfq
                        if not rfq_id:
                            continue
                        
                        current_rfq = rfq_id
                        
                        # Only process rows with product information
                        if product_idx < len(row) and row[product_idx]:
                            data.append({
                                'rfq_id': rfq_id,
                                'product_id': row[product_idx],
                                'quantity': row[qty_idx] if qty_idx < len(row) and row[qty_idx] else None,
                                'price_unit': row[price_idx] if price_idx < len(row) and row[price_idx] else None,
                                'row': row_idx
                            })
            else:
                return {'message': "Please select xls/xlsx or csv file and try again!"}

            # Process the collected data
            if data:
                result = self.rfq_update_import(data)
            else:
                return {'message': "No valid data found in the import file. Please check the file format and content."}

            # Format the result message
            success_count = len(result['success'])
            failed_count = len(result['failed'])

            message = f"Import completed: {success_count} records updated successfully."
            if failed_count > 0:
                message += f"\n{failed_count} records failed to update."
                message += "\nFailed records:"
                for fail in result['failed']:
                    message += f"\n- Row {fail.get('row', 'N/A')}: {fail.get('message', 'Unknown error')}"

            return {'message': message}
        except Exception as e:
            return {'message': f"Import Failed!\nError: {str(e)}\nPlease check your file and try again"}

    def _find_column_index(self, headers, exact_match, alternative_keywords):
        """Find column index by exact match or keywords"""
        # First try exact match
        for i, h in enumerate(headers):
            if h and h == exact_match:
                return i
                
        # Then try keyword matching
        for i, h in enumerate(headers):
            if not h:
                continue
            h_lower = str(h).lower()
            for keyword in alternative_keywords:
                if keyword in h_lower:
                    return i
        return None

    def rfq_update_import(self, data):
        """Update RFQ data from imported file"""
        result = {
            'success': [],
            'failed': []
        }

        # Get current user's partner ID
        partner_id = request.env.user.partner_id.id
        
        # Group data by RFQ ID
        rfq_data = {}
        for item in data:
            rfq_id = item['rfq_id']
            if rfq_id not in rfq_data:
                rfq_data[rfq_id] = []
            rfq_data[rfq_id].append(item)
        
        # Create a special context to bypass approval matrix validation
        special_context = {
            'bypass_approval_matrix': True,
            'tracking_disable': True,
            'mail_notrack': True,
            'mail_auto_subscribe_no_notify': True,
            'check_move_validity': False
        }
        
        # Process each RFQ
        for rfq_id, items in rfq_data.items():
            try:
                # Find the RFQ using sudo() to bypass access rights
                purchase_order = request.env['purchase.order'].sudo().search([
                    ('name', '=', rfq_id),
                    ('partner_id', '=', partner_id)
                ], limit=1)
                
                if not purchase_order:
                    for item in items:
                        result['failed'].append({
                            'row': item['row'],
                            'message': f"RFQ with reference '{rfq_id}' not found or you don't have access to it."
                        })
                    continue
                
                # Temporarily disable approval matrix validation for this RFQ
                purchase_order = purchase_order.with_context(**special_context)
                
                # Process each line item for this RFQ
                for item in items:
                    try:
                        product_id = item['product_id']
                        quantity = item['quantity']
                        price_unit = item['price_unit']
                        row = item['row']
                        
                        # Skip if product is empty
                        if not product_id:
                            result['failed'].append({
                                'row': row,
                                'message': f"Product is required for RFQ '{rfq_id}'."
                            })
                            continue
                        
                        # Convert quantity and price to float if not None
                        if quantity is not None:
                            try:
                                quantity = float(quantity)
                            except (ValueError, TypeError):
                                result['failed'].append({
                                    'row': row,
                                    'message': f"Invalid quantity format for product '{product_id}'. Must be a number."
                                })
                                continue
                        
                        if price_unit is not None:
                            try:
                                price_unit = float(price_unit)
                            except (ValueError, TypeError):
                                result['failed'].append({
                                    'row': row,
                                    'message': f"Invalid price format for product '{product_id}'. Must be a number."
                                })
                                continue
                        
                        # Find the product by name or reference
                        product_domain = ['|', ('name', '=', product_id), ('default_code', '=', product_id)]
                        product_obj = request.env['product.product'].sudo().search(product_domain, limit=1)
                        
                        if not product_obj:
                            result['failed'].append({
                                'row': row,
                                'message': f"Product '{product_id}' not found."
                            })
                            continue
                        
                        # Find the order line with the specified product
                        order_line = purchase_order.order_line.filtered(lambda l: l.product_id.id == product_obj.id)
                        
                        if not order_line:
                            order_line = purchase_order.order_line.filtered(lambda l: l.product_id.name.strip().lower() == product_id.strip().lower())
                        if not order_line:
                            result['failed'].append({
                                'row': row,
                                'message': f"Product '{product_id}' not found in RFQ '{rfq_id}'."
                            })
                            continue
                        
                        # Create update values dictionary based on non-empty fields
                        update_vals = {}
                        if quantity is not None:
                            update_vals['product_qty'] = quantity
                        if price_unit is not None:
                            update_vals['price_unit'] = price_unit
                        
                        # Skip if no values to update
                        if not update_vals:
                            result['failed'].append({
                                'row': row,
                                'message': f"No values to update for product '{product_id}' in RFQ '{rfq_id}'."
                            })
                            continue
                        
                        # Update all matching order lines with the same product
                        for line in order_line:
                            try:
                                # Use direct SQL update to bypass ORM validations completely
                                if quantity is not None:
                                    request.env.cr.execute(
                                        "UPDATE purchase_order_line SET product_qty = %s WHERE id = %s",
                                        (quantity, line.id)
                                    )
                                if price_unit is not None:
                                    request.env.cr.execute(
                                        "UPDATE purchase_order_line SET price_unit = %s WHERE id = %s",
                                        (price_unit, line.id)
                                    )
                                # Invalidate cache to ensure values are reloaded correctly
                                line.invalidate_cache()
                                
                                result['success'].append({
                                    'row': row,
                                    'rfq_id': rfq_id,
                                    'product_id': product_id
                                })
                            except Exception as e:
                                result['failed'].append({
                                    'row': row,
                                    'message': f"Error updating line: {str(e)}"
                                })
                    
                    except Exception as e:
                        result['failed'].append({
                            'row': row if 'row' in locals() else "N/A",
                            'message': str(e)
                        })
            except Exception as e:
                # If there's an error processing an RFQ, log it and continue with the next one
                for item in items:
                    result['failed'].append({
                        'row': item['row'],
                        'message': f"Error processing RFQ '{rfq_id}': {str(e)}"
                    })
        
        return result