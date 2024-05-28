import base64
from tempfile import TemporaryFile

import openpyxl


class ExcelFileReader:
    def __init__(self, file):
        self.file_path = self.generate_excel_file(file)

    def __del__(self):
        self.file_path.close()

    def get_data(self):
        """
        Get the data from the excel file
        :return:  list of data
        """
        data = []
        try:
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active
            for row in sheet.rows:
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                data.append(row_data)
        except Exception as e:
            print(e)
        return data

    def get_data_by_header(self, header):
        """
        Get the data from the excel file
        :param header:  header name
        :return:  list of data
        """
        data = self.get_data()
        column_index = -1
        row_index = -1
        for row in data:
            for column in row:
                if column == header:
                    column_index = row.index(column)
                    row_index = data.index(row)
                    break
        if column_index != -1:
            data = [row[column_index] for row in data[row_index + 1:]]
        return data

    def get_data_by_cell(self, row, column):
        """
        Get the data from the excel file
        :param row:  row number
        :param column:  column number
        :return:  list of data
        """
        data = []
        try:
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active
            data = sheet.cell(row=row, column=column).value
        except Exception as e:
            print(e)
        return data

    @staticmethod
    def generate_excel_file(file):
        """
        Generating of the excel file to be read by openpyxl
        :param file:  base64 encoded file
        :return:  TemporaryFile
        """
        file = base64.decodebytes(file)
        temp_file = TemporaryFile('wb+')
        temp_file.write(file)
        temp_file.seek(0)
        return temp_file
